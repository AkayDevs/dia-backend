from typing import List, Optional, Tuple
import fitz  # PyMuPDF
from pathlib import Path
import os
import shutil
from PIL import Image, ImageDraw, ImageFont
import io
import logging
from fastapi import HTTPException, status
from concurrent.futures import ThreadPoolExecutor
import tempfile
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from docx import Document as DocxDocument
from docx2pdf import convert
import pdf2image

from app.core.config import settings
from app.enums.document import DocumentType
from app.schemas.document import DocumentPage, DocumentPages

logger = logging.getLogger(__name__)

def _get_document_directory(user_id: str, document_id: str) -> Path:
    """Get the document's dedicated directory."""
    return Path(settings.UPLOAD_DIR) / str(user_id) / document_id

def _validate_document_path(document_path: str, user_id: str) -> Path:
    """Validate and return the full document path."""
    try:
        # Clean the path to prevent directory traversal
        clean_path = document_path.replace("/uploads/", "").lstrip("/")
        full_path = Path(settings.UPLOAD_DIR) / clean_path
        
        if not full_path.exists():
            logger.error(f"Document not found: {full_path}")
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Document file not found"
            )
        
        if not full_path.is_file():
            logger.error(f"Path is not a file: {full_path}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid document path"
            )
            
        return full_path
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error validating document path: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error validating document path"
        )

def _setup_document_directories(user_id: str, document_id: str) -> Tuple[Path, Path]:
    """Set up and return the document's original and pages directories."""
    try:
        doc_dir = _get_document_directory(user_id, document_id)
        original_dir = doc_dir / "original"
        pages_dir = doc_dir / "pages"
        
        original_dir.mkdir(parents=True, exist_ok=True)
        pages_dir.mkdir(parents=True, exist_ok=True)
        
        return original_dir, pages_dir
    except Exception as e:
        logger.error(f"Error creating document directories: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error creating document directories"
        )

def _save_image_with_optimization(
    image: Image.Image,
    output_path: Path,
    optimize: bool = True,
    dpi: int = 300
) -> Tuple[int, int]:
    """Save image with optimizations and return dimensions."""
    try:
        # Convert RGBA to RGB if needed
        if image.mode == 'RGBA':
            image = image.convert('RGB')
        
        # Save with optimizations
        image.save(str(output_path), 'PNG', optimize=optimize, dpi=(dpi, dpi))
        return image.size
    except Exception as e:
        logger.error(f"Error saving image: {str(e)}")
        raise

def _create_document_page(
    page_number: int,
    width: int,
    height: int,
    filename: str,
    user_id: str,
    document_id: str
) -> DocumentPage:
    """Create a DocumentPage object with consistent URL formatting."""
    return DocumentPage(
        page_number=page_number,
        width=width,
        height=height,
        image_url=f"/uploads/{user_id}/{document_id}/pages/{filename}"
    )

def _process_pdf_page(args: tuple) -> DocumentPage:
    """Process a single PDF page (for parallel processing)."""
    doc, page_num, pages_dir, user_id, document_id = args
    try:
        page = doc[page_num]
        # Use matrix for better quality (2x zoom)
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
        
        image_filename = f"page_{page_num + 1}.png"
        image_path = pages_dir / image_filename
        
        # Save with error handling
        pix.save(str(image_path))
        
        return _create_document_page(
            page_number=page_num + 1,
            width=pix.width,
            height=pix.height,
            filename=image_filename,
            user_id=user_id,
            document_id=document_id
        )
    except Exception as e:
        logger.error(f"Error processing PDF page {page_num + 1}: {str(e)}")
        raise

def _process_xlsx_to_images(xlsx_path: Path, pages_dir: Path, user_id: str, document_id: str) -> List[DocumentPage]:
    """Convert Excel sheets to images."""
    try:
        pages = []
        workbook = load_workbook(xlsx_path)
        
        for idx, sheet_name in enumerate(workbook.sheetnames):
            ws = workbook[sheet_name]
            
            # Create figure with appropriate size
            plt.figure(figsize=(20, 15))  # Standardized size for better consistency
            
            # Convert sheet data to pandas DataFrame for better visualization
            data = [[cell.value for cell in row] for row in ws.iter_rows()]
            df = pd.DataFrame(data[1:], columns=data[0] if data else None)
            
            # Create and style the table
            ax = plt.subplot(111, frame_on=False)
            ax.xaxis.set_visible(False)
            ax.yaxis.set_visible(False)
            
            table = plt.table(
                cellText=df.values,
                colLabels=df.columns,
                cellLoc='center',
                loc='center',
                bbox=[0, 0, 1, 1]
            )
            
            # Style the table
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1.2, 1.5)
            
            # Save as image
            image_filename = f"sheet_{idx + 1}.png"
            image_path = pages_dir / image_filename
            
            plt.savefig(str(image_path), bbox_inches='tight', dpi=300)
            plt.close()
            
            # Get dimensions and create page object
            with Image.open(image_path) as img:
                width, height = img.size
                
            pages.append(_create_document_page(
                page_number=idx + 1,
                width=width,
                height=height,
                filename=image_filename,
                user_id=user_id,
                document_id=document_id
            ))
            
        return pages
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to process Excel file"
        )

def _process_docx_to_images(docx_path: Path, pages_dir: Path, user_id: str, document_id: str) -> List[DocumentPage]:
    """Convert DOC/DOCX file to images by first converting to PDF."""
    try:
        # Validate file exists and is not empty
        if not docx_path.exists() or docx_path.stat().st_size == 0:
            raise ValueError("File is empty or does not exist")

        # Check if LibreOffice is installed
        import subprocess
        try:
            process = subprocess.run(['soffice', '--version'], capture_output=True, text=True)
            if process.returncode != 0:
                logger.error("LibreOffice is not properly installed")
                raise ValueError("Document conversion software (LibreOffice) is not properly installed")
            logger.info(f"Found LibreOffice: {process.stdout.strip()}")
        except FileNotFoundError:
            logger.error("LibreOffice (soffice) command not found")
            raise ValueError("Document conversion software (LibreOffice) is not installed")

        # Create a temporary directory for PDF conversion
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_pdf = Path(temp_dir) / "temp.pdf"
            
            # Convert DOC/DOCX to PDF using LibreOffice
            process = subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(temp_dir),
                str(docx_path)
            ], capture_output=True, text=True)
            
            if process.returncode != 0:
                logger.error(f"Document conversion failed: {process.stderr}")
                raise ValueError("Failed to convert document to PDF")

            # Find the converted PDF file (it might have a different name)
            pdf_files = list(Path(temp_dir).glob("*.pdf"))
            if not pdf_files:
                raise ValueError("PDF conversion failed - no output file found")
            temp_pdf = pdf_files[0]

            # Now process the PDF using our existing PDF processing code
            try:
                doc = fitz.open(str(temp_pdf))
                try:
                    # Process pages in parallel for better performance
                    with ThreadPoolExecutor() as executor:
                        args = [(doc, i, pages_dir, user_id, document_id) 
                               for i in range(len(doc))]
                        pages = list(executor.map(_process_pdf_page, args))
                finally:
                    doc.close()
                return pages
            except Exception as e:
                logger.error(f"Error processing converted PDF: {str(e)}")
                raise ValueError(f"Error processing converted PDF: {str(e)}")

    except ValueError as ve:
        logger.error(f"Error processing document: {str(ve)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(ve)
        )
    except Exception as e:
        logger.error(f"Error processing document file: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process document file: {str(e)}"
        )

def _get_existing_pages(pages_dir: Path, user_id: str, document_id: str) -> Optional[List[DocumentPage]]:
    """Check if pages already exist for the document and return them."""
    try:
        if not pages_dir.exists():
            return None
            
        # Get all PNG files in the pages directory
        page_files = sorted(pages_dir.glob("*.png"), key=lambda x: int(x.stem.split('_')[1]))
        
        if not page_files:
            return None
            
        pages = []
        for idx, page_file in enumerate(page_files, 1):
            # Get image dimensions
            with Image.open(page_file) as img:
                width, height = img.size
                
            pages.append(_create_document_page(
                page_number=idx,
                width=width,
                height=height,
                filename=page_file.name,
                user_id=user_id,
                document_id=document_id
            ))
            
        return pages if pages else None
        
    except Exception as e:
        logger.warning(f"Error checking existing pages: {str(e)}")
        return None

async def extract_document_pages(document_path: str, document_type: DocumentType, user_id: str, document_id: str) -> DocumentPages:
    """
    Extract pages from a document and convert them to images.
    Supports PDF, DOCX, XLSX, and image files.
    
    Args:
        document_path: Path to the document relative to uploads directory
        document_type: Type of the document (PDF, IMAGE, DOCX, XLSX)
        user_id: ID of the user who owns the document
        document_id: ID of the document being processed
        
    Returns:
        DocumentPages object containing page information and total count
        
    Raises:
        HTTPException: For various error conditions with appropriate status codes
    """
    try:
        # Validate inputs
        if not document_path or not user_id or not document_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Missing required parameters"
            )
        
        # Get validated paths
        full_path = _validate_document_path(document_path, user_id)
        _, pages_dir = _setup_document_directories(user_id, document_id)
        
        # Check for existing pages
        existing_pages = _get_existing_pages(pages_dir, user_id, document_id)
        if existing_pages:
            logger.info(f"Using existing pages for document {document_id}")
            return DocumentPages(
                total_pages=len(existing_pages),
                pages=existing_pages
            )
            
        # If no existing pages, extract them
        pages = []
        
        if document_type == DocumentType.PDF:
            try:
                doc = fitz.open(str(full_path))
                try:
                    # Process pages in parallel for better performance
                    with ThreadPoolExecutor() as executor:
                        args = [(doc, i, pages_dir, user_id, document_id) 
                               for i in range(len(doc))]
                        pages = list(executor.map(_process_pdf_page, args))
                finally:
                    doc.close()
            except fitz.FileDataError as e:
                logger.error(f"Invalid or corrupted PDF file: {str(e)}")
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid or corrupted PDF file"
                )
                
        elif document_type == DocumentType.IMAGE:
            try:
                with Image.open(full_path) as img:
                    image_filename = "page_1.png"
                    image_path = pages_dir / image_filename
                    
                    width, height = _save_image_with_optimization(img, image_path)
                    
                    pages.append(_create_document_page(
                        page_number=1,
                        width=width,
                        height=height,
                        filename=image_filename,
                        user_id=user_id,
                        document_id=document_id
                    ))
            except (IOError, OSError) as e:
                logger.error(f"Error processing image file: {str(e)}")
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid or corrupted image file"
                )
                
        elif document_type == DocumentType.XLSX:
            pages = _process_xlsx_to_images(full_path, pages_dir, user_id, document_id)
            
        elif document_type == DocumentType.DOCX:
            pages = _process_docx_to_images(full_path, pages_dir, user_id, document_id)
            
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unsupported document type: {document_type}"
            )
            
        if not pages:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No pages could be extracted from the document"
            )
            
        return DocumentPages(
            total_pages=len(pages),
            pages=sorted(pages, key=lambda x: x.page_number)
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error extracting document pages: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to extract document pages"
        ) 